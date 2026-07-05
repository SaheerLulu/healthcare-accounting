"""Shared CSV / XLSX HttpResponse builders for register-style report exports.

Every register endpoint supports ?export=csv (and xlsx where noted) — these
helpers keep the response headers and workbook styling identical across the
GST registers (gst_returns) and the books registers (reports).
"""
import csv
import io

from django.http import HttpResponse

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def csv_response(filename, columns, rows, preamble=None):
    """Build a text/csv attachment. `rows` is an iterable of value lists;
    `preamble` is an optional list of single-cell caption lines above the
    header row (period, store, GSTIN...)."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    for line in preamble or []:
        writer.writerow([line])
    if preamble:
        writer.writerow([])
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return response


def xlsx_response(filename, sheets):
    """Build an .xlsx attachment.

    `sheets` is a list of (title, preamble_lines, columns, rows) tuples —
    one worksheet each, with the column header row bolded. Sheet titles are
    clipped to Excel's 31-char limit.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for title, preamble, columns, rows in sheets:
        ws = wb.create_sheet(title=str(title)[:31])
        for line in preamble or []:
            ws.append([line])
        if preamble:
            ws.append([])
        ws.append(list(columns))
        for cell in ws[ws.max_row]:
            cell.font = bold
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type=XLSX_CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
